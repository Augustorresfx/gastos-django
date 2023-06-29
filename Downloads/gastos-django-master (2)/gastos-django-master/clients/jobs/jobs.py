from django.conf import settings
import os
from io import BytesIO
import smtplib
from django.contrib.auth.models import User
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from datetime import timedelta, datetime
from ..models import Operacion, Piloto, Aeronave, Mecanico
from openpyxl import load_workbook
from pathlib import Path
from datetime import date
from django.utils import timezone
from openpyxl.utils import get_column_letter
from django.db.models import Q
load_dotenv()
def schedule_api():

    BASE_DIR = Path(__file__).resolve().parent.parent
    module_dir = os.path.dirname(__file__)   #get current directory
    file_path = os.path.join(BASE_DIR, 'static/Reporte_diario.xlsx')   #full path to text.
    
    path = file_path
    wb = load_workbook(path)
    wb.iso_dates = True
    sheet = wb.active
    max_col = sheet.max_column
    hoy = timezone.now().date()  # Obtiene la fecha actual
    objetos_hoy = Operacion.objects.filter(created__date=hoy)

    objetos_todos = Operacion.objects.all()

    row = 8  # empezar a agregar en la fila 8
    col = 1  # agregar en la primera columna
    for product in objetos_todos:
        timeformat = "%H:%M:%S"
        alumnos = Operacion.objects.filter(Q(alumn__rol__title='alumno'))
    
        instructores = Operacion.objects.filter(Q(alumn__rol__title='instructor'))
        delta = datetime.strptime(str(product.landing_time), timeformat) - datetime.strptime(str(product.takeoff_time), timeformat)
        combustible_usado = product.fuel - product.fuel_on_landing
        data = []
        if product.alumn == None:
            data = [product.created.strftime("%Y"), product.created.strftime("%m"), product.created.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        else:
            if product.alumn.rol.title == 'Copiloto':

            #data = [product.title.title, product.takeoff_place, product.landing_place, product.created.strftime("%y %m %d"), product.pilot.name, product.mechanic.name, product.operator.name, product.aeronave.matricula, timedelta(seconds=delta.seconds), product.reason_of_flight.title, '', '', product.aeronave.horas_voladas, product.start_up_cycles, product.aeronave.horas_disponibles, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.engine_ignition_2, product.engine_cut_2, product.total_encendido_2, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.number_of_landings, product.number_of_splashdowns, product.water_release_cycles, product.water_release_amount]
                data = [product.created.strftime("%Y"), product.created.strftime("%m"), product.created.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), product.alumn.name + ": " + str(product.total_horas_alumn), product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]

            elif product.alumn.rol.title == 'Instructor':
                data = [product.created.strftime("%Y"), product.created.strftime("%m"), product.created.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, product.alumn.name + ": " + str(product.total_horas_alumn), timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
            
            else:
                data = [product.created.strftime("%Y"), product.created.strftime("%m"), product.created.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.aeronave.horas_voladas, product.aeronave.ciclos_motor, product.aeronave.horas_disponibles, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        for i, val in enumerate(data):
            sheet.cell(row=row, column=col+i, value=val)
        row+=1

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    msg = MIMEMultipart()
    msg['From'] = 'no.reply.wings@gmail.com'
    msg['To'] = ', '.join(['augustorresfx@gmail.com',])
    msg['Subject'] = 'Su reporte del día'

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((excel_file).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename='Reporte_Diario' + \
    str(datetime.now())+'.xlsx')
    msg.attach(part)

    html = """\
        <html>
        <head></head>
        <body>
            <h1>Estimado/a,</h1>
            <h2>Adjunto encontrará el archivo de Excel con los datos solicitados:</h2>
        </body>
        </html>
        """
    msg.attach(MIMEText(html, 'html'))

    # Conectar y enviar el correo electrónico
    smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
    smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
    smtp_user = os.getenv('SMTP_USER')
    smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
    smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
    smtp_connection.starttls()
    smtp_connection.login(smtp_user, smtp_password)
    smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
    smtp_connection.quit()



def expiracion_pilotos():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana
    pilotos = Piloto.objects.all()
    expiration_date = date.today() + timedelta(days=30)
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]
    
        # Envía el correo electrónico de recordatorio a cada objeto
    for piloto in pilotos:
        fecha_piloto = piloto.expiration
        fecha_piloto_nuevo_formato = fecha_piloto.strftime('%d/%m/%y')
        if piloto.expiration <= expiration_date and piloto.expiration >= date.today():
            for email in emails_staff:
                    
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(piloto.name)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El psicofísico del piloto: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(piloto.name, fecha_piloto_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

# AERONAVES VENCIMIENTOS FECHAS

def expiracion_aeronaves():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    expiration_date = date.today() + timedelta(days=30)
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        fecha_expiration = aeronave.expiration
        fecha_expiration_nuevo_formato = fecha_expiration.strftime('%d/%m/%y')
        if aeronave.expiration <= expiration_date and aeronave.expiration >= date.today():
            for email in emails_staff:
                    
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El 337 de la aeronave: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, fecha_expiration_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_aeronaves_anexo_2():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    expiration_date = date.today() + timedelta(days=30)
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]       
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        fecha_anexo_2 = aeronave.vencimiento_anexo_2
        fecha_anexo_2_nuevo_formato = fecha_anexo_2.strftime('%d/%m/%y')
        if aeronave.vencimiento_anexo_2 <= expiration_date and aeronave.vencimiento_anexo_2 >= date.today():
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El alquiler anexo 2 de la aeronave: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, fecha_anexo_2_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_aeronaves_notaciones_requerimiento():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    expiration_date = date.today() + timedelta(days=30)
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        fecha_notaciones_requerimiento = aeronave.vencimiento_notaciones_requerimiento
        fecha_notaciones_requerimiento_nuevo_formato = fecha_notaciones_requerimiento.strftime('%d/%m/%y')
        if aeronave.vencimiento_notaciones_requerimiento <= expiration_date and aeronave.vencimiento_notaciones_requerimiento >= date.today():
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>Notaciones de requerimiento de la aeronave: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, fecha_notaciones_requerimiento_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_aeronaves_inspeccion_anual():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    expiration_date = date.today() + timedelta(days=30)
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]        
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        fecha_inspeccion_anual = aeronave.vencimiento_inspeccion_anual
        fecha_inspeccion_anual_nuevo_formato = fecha_inspeccion_anual.strftime('%d/%m/%y')
        if aeronave.vencimiento_inspeccion_anual <= expiration_date and aeronave.vencimiento_inspeccion_anual >= date.today():
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El vencimiento de inspección anual de la aeronave: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, fecha_inspeccion_anual_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

# AERONAVES VENCIMIENTOS HORAS

def expiracion_aeronaves_25():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    hora_proxima = 6.0 
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]        
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        if aeronave.horas_inspecciones_varias_25 <= hora_proxima:
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El remanente de inspecciones varias (25hs) de la aeronave: {}</h2>
                        <h3>Está proximo a llegar a cero, con horas: {}</h3>
                        <h3>Se recomienda actualizar la cantidad de horas para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, aeronave.horas_inspecciones_varias_25)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_aeronaves_50():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    hora_proxima = 6.0 
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]        
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        if aeronave.horas_inspecciones_varias_50 <= hora_proxima:
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El remanente de inspecciones varias (50hs) de la aeronave: {}</h2>
                        <h3>Está proximo a llegar a cero, con horas: {}</h3>
                        <h3>Se recomienda actualizar la cantidad de horas para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, aeronave.horas_inspecciones_varias_50)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_aeronaves_100():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    hora_proxima = 6.0 
    aeronaves = Aeronave.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]        
        # Envía el correo electrónico de recordatorio a cada objeto
    for aeronave in aeronaves:
        if aeronave.horas_inspecciones_varias_100 <= hora_proxima:
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(aeronave.title)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El remanente de inspecciones varias (100hs) de la aeronave: {}</h2>
                        <h3>Está proximo a llegar a cero, con horas: {}</h3>
                        <h3>Se recomienda actualizar la cantidad de horas para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(aeronave.title, aeronave.horas_inspecciones_varias_100)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

def expiracion_mecanicos():
    # Obtén los objetos cuya fecha de expiración es dentro de una semana

    expiration_date = date.today() + timedelta(days=30)
    mecanicos = Mecanico.objects.all()
    usuarios_staff = User.objects.filter(is_staff=True)
    emails_staff = [user.email for user in usuarios_staff]
        # Envía el correo electrónico de recordatorio a cada objeto
    for mecanico in mecanicos:
        fecha_mecanico = mecanico.expiration
        fecha_mecanico_nuevo_formato = fecha_mecanico.strftime('%d/%m/%y')
        if mecanico.expiration <= expiration_date and mecanico.expiration >= date.today():
            for email in emails_staff:
                msg = MIMEMultipart()
                msg['From'] = 'no.reply.wings@gmail.com'
                msg['To'] = email
                msg['Subject'] = 'Notificación vencimiento: {}'.format(mecanico.name)

                html = """\
                    <html>
                    <head></head>
                    <body>
                        <h1>Estimado/a,</h1>
                        <h2>El psicofísico del mecánico: {}</h2>
                        <h3>Está proximo a expirar, con fecha: {}</h3>
                        <h3>Se recomienda actualizar la fecha para evitar errores, muchas gracias.</h3>
                    </body>
                    </html>
                    """.format(mecanico.name, fecha_mecanico_nuevo_formato)
                msg.attach(MIMEText(html, 'html'))
                    # Conectar y enviar el correo electrónico
                smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
                smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
                smtp_user = os.getenv('SMTP_USER')
                smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
                smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
                smtp_connection.starttls()
                smtp_connection.login(smtp_user, smtp_password)
                smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
                smtp_connection.quit()

