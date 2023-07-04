from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from .forms import ClientForm, ClientFilterForm, GastoForm, AeronaveForm, OtroForm, PilotoForm, MecanicoForm
from .models import Operacion, Aeronave, Mecanico, Piloto, Otro
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from .filters import ProductFilter, GastosFilter
from django.http import HttpResponse
import xlwt
from datetime import timedelta, datetime, date, time
from django.http import JsonResponse
import json
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from .models import Gasto
import os
from io import BytesIO
import smtplib
from email.mime.text import MIMEText
from django.db.models import Q

from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

from dotenv import load_dotenv
from django.utils.formats import date_format
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.urls import reverse_lazy

def is_staff_user(user):
    return user.is_staff

staff_required = user_passes_test(is_staff_user, login_url=reverse_lazy('signin'))


def is_basic_user(user):
    return user.groups.filter(name='PermisoBasico').exists()
load_dotenv()
# Create your views here.
def home(request):
    return render(request, 'home.html')

# def signup(request):

#     if request.method == 'GET':
#         return render(request, 'signup.html', {
#             'form': UserCreationForm
#         })
#     else:
#         if request.POST['password1'] == request.POST['password2']:
#             try:
              
#                 user = User.objects.create_user(username=request.POST['username'], 
#                 password=request.POST['password1'])
#                 user.save()
#                 login(request, user)
#                 return redirect('vuelos')
#             except IntegrityError:
#                 return render(request, 'signup.html', {
#                     'form': UserCreationForm,
#                     'error': 'Usuario ya existe'
#                 })
#         else:
#             return render(request, 'signup.html', {
#                 'form': UserCreationForm,
#                 'error': 'Las constraseñas no coinciden'
#             })

def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        expenses = Operacion.objects.filter(
            price__istartswith=search_str, ) | Operacion.objects.filter(
            created__istartswith=search_str, ) | Operacion.objects.filter(
            description__icontains=search_str, ) | Operacion.objects.filter(
            category__icontains=search_str, )
        data = expenses.values()
        return JsonResponse(list(data), safe=False)

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def vuelos(request):
  
    context = {}   
    
    filtered_clients = ProductFilter(
        request.GET, 
        queryset=Operacion.objects.all().order_by('-created'),

        )

    context['clients'] = filtered_clients.qs
    paginated_filter = Paginator(filtered_clients.qs, 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
   
    context['form'] = filtered_clients.form
    context['pages'] = filter_pages
    
    return render(request, 'vuelos/vuelos.html', context=context)

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def expensas(request):
  
    context = {}
    filtered_clients = GastosFilter(
        request.GET, 
        queryset=Gasto.objects.all().order_by('-id'),

        )
    context['clients'] = filtered_clients
    paginated_filter = Paginator(Gasto.objects.all(), 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
    context['pages'] = filter_pages
   
    return render(request, 'expensas/expensas.html', context=context)

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def create_expensa(request):
    if request.method == 'GET':

        return render(request, 'expensas/create_expensa.html', {
            'form': GastoForm,
        })
    else:
        try:
            
            form = GastoForm(request.POST)
            new_client = form.save(commit=False)
            new_client.user = request.user
            new_client.save()
            return redirect('expensas')
        except ValueError:
            return render(request, 'expensas/create_expensa.html', {
                'form': GastoForm,
                'error': 'Please provide valid data'
            })

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def create_vuelo(request):
    if request.method == 'GET':

        return render(request, 'vuelos/create_vuelo.html', {
            'form': ClientForm,

        })
    else:
        try:
            
            form = ClientForm(request.POST)
            new_client = form.save(commit=False)
            hora_despuegue = new_client.takeoff_time
            hora_aterrizaje = new_client.landing_time
            hora_encendido = new_client.engine_ignition_1
            hora_corte = new_client.engine_cut_1
            if hora_aterrizaje < hora_despuegue:
                raise ValueError("La hora de aterrizaje no puede ser anterior a la hora de despegue.")
            if hora_corte < hora_encendido:
                raise ValueError("La hora de puesta en marcha no puede ser anterior a la hora de corte de motor.")
            new_client.user = request.user
            new_client.save()
            return redirect('vuelos')
        except ValueError as e:
            if str(e) == "La hora de aterrizaje no puede ser anterior a la hora de despegue.":
                return render(request, 'vuelos/create_vuelo.html', {
                    'form': ClientForm,
                    'error': str(e)
                })
            elif str(e) == "La hora de puesta en marcha no puede ser anterior a la hora de corte de motor.":
                return render(request, 'vuelos/create_vuelo.html', {
                    'form': ClientForm,
                    'error': str(e)
                })
            else:
                return render(request, 'vuelos/create_vuelo.html', {
                    'form': ClientForm,
                    'error': 'Please provide valid data.'
                })

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def expensa_detail(request, gasto_id):
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    form = GastoForm(instance=gasto)
    if request.method == 'GET':
        if gasto.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para ver este elemento')
            return redirect('expensas')
        
        if (gasto.subtotal):
            subtotal_formatted = "{:.2f}".format(gasto.subtotal).replace(',', '.')
        else:
            subtotal_formatted = 0
        if (gasto.iva_total):
            iva_total_formatted = "{:.2f}".format(gasto.iva_total).replace(',', '.')
        else:
            iva_total_formatted = 0
        if (gasto.concepto_no_grabado_total):
            concepto_no_grabado_total_formatted = "{:.2f}".format(gasto.concepto_no_grabado_total).replace(',', '.')
        else:
            concepto_no_grabado_total_formatted = 0
        if (gasto.impuesto_vario_total):
            impuesto_vario_total_formatted = "{:.2f}".format(gasto.impuesto_vario_total).replace(',', '.')
        else:
            impuesto_vario_total_formatted = 0
        return render(request, 'expensas/expensa_detail.html', {'expensa': gasto, 'form': form, 'subtotal_formatted': subtotal_formatted, 'iva_total_formatted': iva_total_formatted, 'concepto_no_grabado_total_formatted': concepto_no_grabado_total_formatted, 'impuesto_vario_total_formatted': impuesto_vario_total_formatted})

    elif request.method == 'POST':
        if gasto.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para editar este elemento')
            return redirect('expensas')
        
        form = GastoForm(request.POST, instance=gasto)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Elemento editado correctamente')
        else:
            messages.error(request, 'Error actualizando la operación')
        
        return redirect('expensas')
    
    # Si llegamos a este punto, se trata de un método no permitido o una solicitud inválida
    return redirect('expensas')
    


@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

# def delete_expensa(request, gasto_id):
#     gasto = get_object_or_404(Gasto, pk=gasto_id)
#     if request.method == 'POST':
#         gasto.delete()
#         return redirect('expensas')

def delete_expensa(request, gasto_id):
    gasto = get_object_or_404(Gasto, pk=gasto_id)
    if gasto.user == request.user or request.user.is_staff:
        if request.method == 'POST':
            gasto.delete()
            messages.success(request, 'Elemento eliminado correctamente')
            return redirect('expensas')
        else:
            return render(request, 'expensas/delete_expensa.html', {'gasto': gasto})
    else:
        messages.error(request, 'No tienes permiso para eliminar este elemento')
        return redirect('expensas')

# EXCEL DE EXPENSAS
@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def expensas_send_mail_with_excel(request):
    module_dir = os.path.dirname(__file__)   #get current directory
    file_path = os.path.join(module_dir, 'static/Reporte_expensas.xlsx')   #full path to text.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_expensas ' + \
    str(timezone.now().date())+'.xlsx'
    path = file_path
    wb = load_workbook(path)
    wb.iso_dates = True
    sheet = wb.active
    max_col = sheet.max_column

    # Obtener fecha y hora actual en la zona horaria del proyecto
    hoy = timezone.now().date()  # Obtiene la fecha actual
    objetos_hoy = Gasto.objects.filter(created__date=hoy)
    objetos_todos = Gasto.objects.all()
    user_email = request.user.email
    row = 8  # empezar a agregar en la fila 8
    col = 1  # agregar en la primera columna
    for product in objetos_todos:
        timeformat = "%H:%M:%S"
       
        data = []
        data.append(product.id)

       

        if product.base:
            data.append(product.base.title)
            data.append('')
            data.append('')

        elif product.aeronave:
            data.append('')
            data.append(product.aeronave.title)
            data.append('')

        elif product.traslado:
            data.append('')
            data.append('')
            data.append(product.traslado.title)

        else:
            data.append('')
            data.append('')
            data.append('')

        if product.responsable:
            data.append(product.responsable.username)
        else:
            data.append('')

        if product.categoria:
            data.append(product.categoria.title)
        else:
            data.append('')
            

        if product.fecha_compra:
            data.append(product.fecha_compra)
        else:
            data.append('')
        if product.numero_compra:
            data.append(product.numero_compra)
        else:
            data.append('')
        if product.cuit:
            data.append(product.cuit)
        else:
            data.append('')


        if product.moneda:
            data.append(product.moneda.representacion)
        else:
            data.append('')

        if product.subtotal:
            data.append(product.subtotal)
        else:
            data.append('0')

        if product.concepto_no_grabado:
            data.append(product.concepto_no_grabado.title)
        else:
            data.append('')

        if product.concepto_no_grabado_total:
            data.append(product.concepto_no_grabado_total)
        else:
            data.append('0')

        if product.iva:
            data.append(product.iva.title)
        else:
            data.append('')

        if product.iva_total:
            data.append(product.iva_total)
        else:
            data.append('0')
      
        if product.impuesto_vario:
            data.append(product.impuesto_vario.title)
        else:
            data.append('')

        if product.impuesto_vario_total:
            data.append(product.impuesto_vario_total)
        else:
            data.append('0')
        

        if product.total:
            data.append(product.total)
        else:
            data.append('0')
        for i, val in enumerate(data):
            sheet.cell(row=row, column=col+i, value=val)
        row+=1

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    destinatarios = ['gdguerra07@gmail.com', 'gguerra@helicopterosdelpacifico.com.ar', 'augustorresfx@gmail.com']
    destinatarios2 = ['augustorresfx@gmail.com']
    try:

        msg = MIMEMultipart()
        msg['From'] = 'no.reply.wings@gmail.com'
        msg['To'] = user_email

        msg['Subject'] = 'Su reporte de expensas'

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((excel_file).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename='Reporte_expensas' + \
        str(datetime.now())+'.xlsx')
        msg.attach(part)

        html = """\
            <html>
            <head></head>
            <body>
                <h1>Estimado/a,</h1>
                <h2>Adjunto encontrará el archivo de Excel con los datos solicitados:</h2>
            </body>
            </html>
            """
        msg.attach(MIMEText(html, 'html'))

        # Conectar y enviar el correo electrónico
        smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
        smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
        smtp_user = os.getenv('SMTP_USER')
        smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
        smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
        smtp_connection.starttls()
        smtp_connection.login(smtp_user, smtp_password)
        smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
        smtp_connection.quit()
        excel_file.seek(0)
        messages.success(request, 'Los correos electrónicos se enviaron correctamente')
    except:
        messages.error(request, 'Error enviando los correos electrónicos')
    
    
    return redirect('expensas')

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def expensas_export_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_expensas ' + \
        str(timezone.now().date())+'.xlsx'
    module_dir = os.path.dirname(__file__)   #get current directory
    file_path = os.path.join(module_dir, 'static/Reporte_expensas.xlsx')   #full path to text.
    path = file_path
    wb = load_workbook(path)
    wb.iso_dates = True
    sheet = wb.active
    max_col = sheet.max_column

    
    hoy = timezone.now().date()  # Obtiene la fecha actual
    objetos_hoy = Gasto.objects.filter(created__date=hoy)

    objetos_todos = Gasto.objects.all()
    row = 8  # empezar a agregar en la fila 8
    col = 1  # agregar en la primera columna
    for product in objetos_todos:
        timeformat = "%H:%M:%S"
       
        data = []
        data.append(product.id)

       

        if product.base:
            data.append(product.base.title)
            data.append('')
            data.append('')

        elif product.aeronave:
            data.append('')
            data.append(product.aeronave.title)
            data.append('')

        elif product.traslado:
            data.append('')
            data.append('')
            data.append(product.traslado.title)

        else:
            data.append('')
            data.append('')
            data.append('')

        if product.responsable:
            data.append(product.responsable.username)
        else:
            data.append('')

        if product.categoria:
            data.append(product.categoria.title)
        else:
            data.append('')
            

        if product.fecha_compra:
            data.append(product.fecha_compra)
        else:
            data.append('')
        if product.numero_compra:
            data.append(product.numero_compra)
        else:
            data.append('')
        if product.cuit:
            data.append(product.cuit)
        else:
            data.append('')


        if product.moneda:
            data.append(product.moneda.representacion)
        else:
            data.append('')

        if product.subtotal:
            data.append(product.subtotal)
        else:
            data.append('0')

        if product.concepto_no_grabado:
            data.append(product.concepto_no_grabado.title)
        else:
            data.append('')

        if product.concepto_no_grabado_total:
            data.append(product.concepto_no_grabado_total)
        else:
            data.append('0')

        if product.iva:
            data.append(product.iva.title)
        else:
            data.append('')

        if product.iva_total:
            data.append(product.iva_total)
        else:
            data.append('0')
      
        if product.impuesto_vario:
            data.append(product.impuesto_vario.title)
        else:
            data.append('')

        if product.impuesto_vario_total:
            data.append(product.impuesto_vario_total)
        else:
            data.append('0')
        

        if product.total:
            data.append(product.total)
        else:
            data.append('0')
        for i, val in enumerate(data):
            sheet.cell(row=row, column=col+i, value=val)
        row+=1


    wb.save(response)  

    return response

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def vuelo_detail(request, product_id):
    product = get_object_or_404(Operacion, pk=product_id)

    if request.method == 'GET':
        if product.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para ver este elemento')
            return redirect('vuelos')
        
        form = ClientForm(instance=product)
        return render(request, 'vuelos/vuelo_detail.html', {'client': product, 'form': form})

    elif request.method == 'POST':
        if product.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para editar este elemento')
            return redirect('vuelos')
        
        form = ClientForm(request.POST, instance=product)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Elemento editado correctamente')
        else:
            messages.error(request, 'Error actualizando la operación')
        
        return redirect('vuelos')
    
    # Si llegamos a este punto, se trata de un método no permitido o una solicitud inválida
    return redirect('vuelos')
    
@login_required
@user_passes_test(lambda user: user.is_staff)

def send_mail_with_excel(request):
    module_dir = os.path.dirname(__file__)   #get current directory
    file_path = os.path.join(module_dir, 'static/Reporte_diario.xlsx')   #full path to text.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Diario ' + \
    str(timezone.now().date())+'.xlsx'
    path = file_path
    wb = load_workbook(path)
    wb.iso_dates = True
    sheet = wb.active
    max_col = sheet.max_column

    # Obtener fecha y hora actual en la zona horaria del proyecto
    hoy = timezone.now().date()  # Obtiene la fecha actual
    objetos_hoy = Operacion.objects.filter(created__date=hoy)
    objetos_todos = Operacion.objects.all().order_by('-created')
    copilotos = Otro.objects.filter(rol__title='copiloto')
    row = 8  # empezar a agregar en la fila 8
    col = 1  # agregar en la primera columna
    user_email = request.user.email
    for product in objetos_todos:
        print(product.fecha)
        timeformat = "%H:%M:%S"
        alumnos = Operacion.objects.filter(Q(alumn__rol__title='alumno'))
    
        instructores = Operacion.objects.filter(Q(alumn__rol__title='instructor'))
        delta = datetime.strptime(str(product.landing_time), timeformat) - datetime.strptime(str(product.takeoff_time), timeformat)
        combustible_usado = product.fuel - product.fuel_on_landing
        data = []
        if product.alumn == None:
            data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        else:
            if product.alumn.rol.title == 'Copiloto':
                data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), product.alumn.name + ": " + str(product.total_horas_alumn), product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
           

            elif product.alumn.rol.title == 'Instructor':
                data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, product.alumn.name + ": " + str(product.total_horas_alumn), timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
            
            else:
                 data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.number_of_landings, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        for i, val in enumerate(data):
            sheet.cell(row=row, column=col+i, value=val)
        row+=1

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    destinatarios = ['gdguerra07@gmail.com', 'gguerra@helicopterosdelpacifico.com.ar', 'augustorresfx@gmail.com']
    destinatarios2 = ['augustorresfx@gmail.com']
    try:
    
        msg = MIMEMultipart()
        msg['From'] = 'no.reply.wings@gmail.com'
        msg['To'] = user_email

        msg['Subject'] = 'Su reporte de operaciones'

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((excel_file).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename='Reporte_Diario' + \
        str(datetime.now())+'.xlsx')
        msg.attach(part)

        html = """\
            <html>
            <head></head>
            <body>
                <h1>Estimado/a,</h1>
                <h2>Adjunto encontrará el archivo de Excel con los datos solicitados:</h2>
            </body>
            </html>
            """
        msg.attach(MIMEText(html, 'html'))

        # Conectar y enviar el correo electrónico
        smtp_server = 'smtp.gmail.com'  # Cambia esto a tu servidor SMTP
        smtp_port = 587  # Cambia esto al puerto de tu servidor SMTP
        smtp_user = os.getenv('SMTP_USER')
        smtp_password = os.getenv('SMTP_PASSWORD')  # Cambia esto a tu contraseña SMTP
        smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
        smtp_connection.starttls()
        smtp_connection.login(smtp_user, smtp_password)
        smtp_connection.sendmail(smtp_user, msg['To'], msg.as_string())
        smtp_connection.quit()
        excel_file.seek(0)
        messages.success(request, 'Los correos electrónicos se enviaron correctamente')
    except:
        messages.error(request, 'Error enviando los correos electrónicos')
    
    
    return redirect('vuelos')

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def export_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Diario ' + \
        str(timezone.now().date())+'.xlsx'
    module_dir = os.path.dirname(__file__)   #get current directory
    file_path = os.path.join(module_dir, 'static/Reporte_diario.xlsx')   #full path to text.
    path = file_path
    wb = load_workbook(path)
    wb.iso_dates = True
    sheet = wb.active
    max_col = sheet.max_column

    
    hoy = timezone.now().date()  # Obtiene la fecha actual
    objetos_hoy = Operacion.objects.filter(created__date=hoy)

    objetos_todos = Operacion.objects.all().order_by('-created')
    copilotos = Otro.objects.filter(rol__title='copiloto')
    row = 8  # empezar a agregar en la fila 8
    col = 1  # agregar en la primera columna
    for product in objetos_todos:
        timeformat = "%H:%M:%S"
        alumnos = Operacion.objects.filter(Q(alumn__rol__title='alumno'))
    
        instructores = Operacion.objects.filter(Q(alumn__rol__title='instructor'))
        delta = datetime.strptime(str(product.landing_time), timeformat) - datetime.strptime(str(product.takeoff_time), timeformat)
        combustible_usado = product.fuel - product.fuel_on_landing
        data = []
        if product.alumn == None:
            data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.total_aterrizajes, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        else:
            if product.alumn.rol.title == 'Copiloto':
                data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), product.alumn.name + ": " + str(product.total_horas_alumn), product.total_aterrizajes, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
           

            elif product.alumn.rol.title == 'Instructor':
                data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.total_aterrizajes, product.number_of_splashdowns, product.alumn.name + ": " + str(product.total_horas_alumn), timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
            
            else:
                 data = [product.created.strftime("%Y/%m/%d"), product.fecha.strftime("%Y"), product.fecha.strftime("%m"), product.fecha.strftime("%d"), product.takeoff_time, product.takeoff_place, product.landing_place, product.landing_time, product.reason_of_flight.title, product.aeronave.title, product.aeronave.matricula, product.pilot.name + ': ' + str(product.total_horas_piloto), '', product.total_aterrizajes, product.number_of_splashdowns, '', timedelta(seconds=delta.seconds), product.reason_of_flight.title, product.total_horas_aeronave, product.total_ciclos_encendido, product.cant_pasajeros, product.total_horas_disponibles_aeronave, '', product.fuel, product.fuel_on_landing, product.used_fuel, product.engine_ignition_1, product.engine_cut_1, product.total_encendido_1, product.operation_note, product.maintenance_note, product.client.name, product.cycles_with_external_load, product.weight_with_external_load, product.water_release_cycles, product.water_release_amount ]
        for i, val in enumerate(data):
            sheet.cell(row=row, column=col+i, value=val)
        row+=1


    wb.save(response)  

    return response

@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)
# def delete_gasto(request, product_id):
#     product = get_object_or_404(Operacion, pk=product_id)
#     if request.method == 'POST':
#         product.delete()
#         return redirect('gastos')


def delete_vuelo(request, product_id):
    product = get_object_or_404(Operacion, pk=product_id)
    if product.user == request.user or request.user.is_staff:
        if request.method == 'POST':
            product.delete()
            messages.success(request, 'Elemento eliminado correctamente')
            return redirect('vuelos')
        else:
            return render(request, 'vuelos/delete_vuelo.html', {'product': product})
    else:
        messages.error(request, 'No tienes permiso para eliminar este elemento')
        return redirect('vuelos')

#  AERONAVES


@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)

def aeronaves(request):

    context = {}   
    context['aeronaves'] = Aeronave.objects.all()
    paginated_filter = Paginator(Aeronave.objects.all(), 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
    context['pages'] = filter_pages
   
    return render(request, 'aeronaves/aeronaves.html', context=context)


@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)


def create_aeronave(request):
    if request.method == 'GET':

        return render(request, 'aeronaves/create_aeronave.html', {
            'form': AeronaveForm,

        })
    else:
        try:
            
            form = AeronaveForm(request.POST)
            new_client = form.save(commit=False)
            new_client.user = request.user
            new_client.save()
            return redirect('aeronaves')
        except ValueError:
            return render(request, 'aeronaves/create_aeronave.html', {
                'form': AeronaveForm,
                'error': 'Please provide valid data'
            })


@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)


def aeronave_detail(request, aeronave_id):
    aeronave = get_object_or_404(Aeronave, pk=aeronave_id)

    if request.method == 'GET':
        if aeronave.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para ver este elemento')
            return redirect('aeronaves')
        
        form = AeronaveForm(instance=aeronave)
        horas_voladas_formatted = "{:.2f}".format(aeronave.horas_voladas).replace(',', '.')
        horas_disponibles_formatted = "{:.2f}".format(aeronave.horas_disponibles).replace(',', '.')
        horas_inspecciones_varias_25_formatted = "{:.2f}".format(aeronave.horas_inspecciones_varias_25).replace(',', '.')
        horas_inspecciones_varias_50_formatted = "{:.2f}".format(aeronave.horas_inspecciones_varias_50).replace(',', '.')
        horas_inspecciones_varias_100_formatted = "{:.2f}".format(aeronave.horas_inspecciones_varias_100).replace(',', '.')
        return render(request, 'aeronaves/aeronave_detail.html', {'aeronave': aeronave, 'form': form, 'horas_voladas_formatted': horas_voladas_formatted, 'horas_disponibles_formatted': horas_disponibles_formatted, 'horas_inspecciones_varias_25_formatted': horas_inspecciones_varias_25_formatted, 'horas_inspecciones_varias_50_formatted': horas_inspecciones_varias_50_formatted, 'horas_inspecciones_varias_100_formatted': horas_inspecciones_varias_100_formatted})
    elif request.method == 'POST':
        if aeronave.user != request.user and not request.user.is_staff:
            messages.error(request, 'No tienes permiso para editar este elemento')
            return redirect('aeronaves')
        
        form = AeronaveForm(request.POST, instance=aeronave)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Elemento editado correctamente')
        else:
            messages.error(request, 'Error actualizando la operación')
        
        return redirect('aeronaves')



@login_required
@user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff)


def delete_aeronave(request, aeronave_id):
    aeronave = get_object_or_404(Operacion, pk=aeronave_id)
    if aeronave.user == request.user or request.user.is_staff:
        if request.method == 'POST':
            aeronave.delete()
            messages.success(request, 'Elemento eliminado correctamente')
            return redirect('aeronaves')
        else:
            return render(request, 'aeronaves/delete_aeronave.html', {'aeronave': aeronave})
    else:
        messages.error(request, 'No tienes permiso para eliminar este elemento')
        return redirect('aeronaves')


# FIN AERONAVES

# PILOTOS

@login_required
@staff_required

def pilotos(request):
  
    context = {}   
    context['pilotos'] = Piloto.objects.all()
    paginated_filter = Paginator(Piloto.objects.all(), 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
    context['pages'] = filter_pages
   
    return render(request, 'pilotos/pilotos.html', context=context)

@login_required
@staff_required

def create_piloto(request):
    if request.method == 'GET':

        return render(request, 'pilotos/create_piloto.html', {
            'form': PilotoForm,

        })
    else:
        try:
            
            form = PilotoForm(request.POST)
            new_client = form.save(commit=False)
            new_client.user = request.user
            new_client.save()
            return redirect('pilotos')
        except ValueError:
            return render(request, 'pilotos/create_piloto.html', {
                'form': PilotoForm,
                'error': 'Please provide valid data'
            })

@login_required
@staff_required

def piloto_detail(request, piloto_id):
    if request.method == 'GET':
        piloto = get_object_or_404(Piloto, pk=piloto_id)
        form = PilotoForm(instance=piloto)
        horas_voladas_formatted = "{:.2f}".format(piloto.horas_voladas).replace(',', '.')
        return render(request, 'pilotos/piloto_detail.html', {'piloto': piloto, 'form': form, 'horas_voladas_formatted': horas_voladas_formatted,})
    else:
        try:
            piloto = get_object_or_404(Piloto, pk=piloto_id)
            form = PilotoForm(request.POST, instance=piloto)
            form.save()
            return redirect('pilotos')
        except ValueError:
            return render(request, 'pilotos/piloto_detail.html', {'piloto': piloto, 'form': form, 'error': 'Error actualizando el piloto'})


@login_required
@staff_required

def delete_piloto(request, piloto_id):
    piloto = get_object_or_404(Piloto, pk=piloto_id)
    if request.method == 'POST':
        piloto.delete()
        return redirect('pilotos')

# FIN PILOTOS

# MECANICOS

@login_required
@staff_required

def mecanicos(request):
  
    context = {}   
    context['mecanicos'] = Mecanico.objects.all()
    paginated_filter = Paginator(Mecanico.objects.all(), 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
    context['pages'] = filter_pages
   
    return render(request, 'mecanicos/mecanicos.html', context=context)

@login_required
@staff_required

def create_mecanico(request):
    if request.method == 'GET':

        return render(request, 'mecanicos/create_mecanico.html', {
            'form': MecanicoForm,

        })
    else:
        try:
            
            form = MecanicoForm(request.POST)
            new_client = form.save(commit=False)
            new_client.user = request.user
            new_client.save()
            return redirect('mecanicos')
        except ValueError:
            return render(request, 'mecanicos/create_mecanico.html', {
                'form': MecanicoForm,
                'error': 'Please provide valid data'
            })

@login_required
@staff_required

def mecanico_detail(request, mecanico_id):
    if request.method == 'GET':
        mecanico = get_object_or_404(Mecanico, pk=mecanico_id)
        form = MecanicoForm(instance=mecanico)
        return render(request, 'mecanicos/mecanico_detail.html', {'mecanico': mecanico, 'form': form})
    else:
        try:
            mecanico = get_object_or_404(Mecanico, pk=mecanico_id)
            form = MecanicoForm(request.POST, instance=mecanico)
            form.save()
            return redirect('mecanicos')
        except ValueError:
            return render(request, 'mecanicos/mecanico_detail.html', {'mecanico': mecanico, 'form': form, 'error': 'Error actualizando el mecánico'})

@login_required
@staff_required

def delete_mecanico(request, mecanico_id):
    mecanico = get_object_or_404(Mecanico, pk=mecanico_id)
    if request.method == 'POST':
        mecanico.delete()
        return redirect('mecanicos')

# FIN MECANICOS

# INICIO OTROS
@login_required
@staff_required

def otros(request):
  
    context = {}   
    context['otros'] = Otro.objects.all()
    paginated_filter = Paginator(Otro.objects.all(), 10)
    page_number = request.GET.get("page")
    filter_pages = paginated_filter.get_page(page_number)
    context['pages'] = filter_pages
   
    return render(request, 'otros/otros.html', context=context)

@login_required
@staff_required

def create_otro(request):
    if request.method == 'GET':

        return render(request, 'otros/create_otro.html', {
            'form': OtroForm,

        })
    else:
        try:
            
            form = OtroForm(request.POST)
            new_client = form.save(commit=False)
            new_client.user = request.user
            new_client.save()
            return redirect('otros')
        except ValueError:
            return render(request, 'otros/create_otro.html', {
                'form': OtroForm,
                'error': 'Please provide valid data'
            })

@login_required
@staff_required

def otro_detail(request, otro_id):
    if request.method == 'GET':
        otro = get_object_or_404(Otro, pk=otro_id)
        horas_voladas_formatted = "{:.2f}".format(otro.horas_voladas).replace(',', '.')
        form = OtroForm(instance=otro)
        return render(request, 'otros/otro_detail.html', {'otro': otro, 'form': form, 'horas_voladas_formatted': horas_voladas_formatted,})
    else:
        try:
            otro = get_object_or_404(Otro, pk=otro_id)
            form = OtroForm(request.POST, instance=otro)
            form.save()
            return redirect('otros')
        except ValueError:
            return render(request, 'otros/otro_detail.html', {'otro': otro, 'form': form, 'error': 'Error actualizando'})


@login_required
@staff_required

def delete_otro(request, otro_id):
    otro = get_object_or_404(Otro, pk=otro_id)
    if request.method == 'POST':
        otro.delete()
        return redirect('otros')

# FIN OTROS
@login_required

def signout(request):
    logout(request)
    return redirect('home')

def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {
            'form': AuthenticationForm
        })
    else:
        user = authenticate(
            request, username=request.POST['username'], 
            password=request.POST['password'])
        if user is None:
            return render(request, 'signin.html', {
            'form': AuthenticationForm,
            'error': 'El nombre de usuario o la contraseña no existen',
            })
        else:
            login(request, user)
            return redirect('vuelos')
